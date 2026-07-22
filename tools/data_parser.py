"""
Data Parser Tools — universal CSV/XLSX/JSON parsers for all agents.

Added to BASE_CAPABILITY_TOOLS so EVERY agent can ingest data files
without needing domain-specific tool registrations.
"""
import csv
import json
import sys
from io import StringIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def _resolve_path(file_path: str) -> Path:
    """Resolve a file path — absolute, relative to CWD, or Windows path."""
    p = Path(file_path)
    if p.is_absolute() and p.exists():
        return p
    # Try relative to CWD
    cwd_path = Path.cwd() / file_path
    if cwd_path.exists():
        return cwd_path
    # The path might be a valid absolute that just doesn't exist yet
    return p


async def parse_spreadsheet(file_path: str) -> str:
    """Parse a CSV or XLSX file and return structured JSON.

    Supports:
    - .csv — comma/tab/semicolon-delimited (auto-detected)
    - .xlsx — Excel (requires openpyxl; degrades gracefully if missing)

    Returns JSON with: {file, row_count, columns, data (first 100 rows), truncated (bool)}
    """
    p = _resolve_path(file_path)

    if not p.exists():
        return json.dumps({
            "error": f"File not found: {file_path}",
            "resolved_path": str(p),
            "hint": "Provide an absolute path or a path relative to the project root",
        }, indent=2)

    suffix = p.suffix.lower()

    # ── CSV ──
    if suffix == ".csv":
        try:
            text = p.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = p.read_text(encoding="latin-1")

        # Auto-detect delimiter
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(text[:4096])
            delimiter = dialect.delimiter
        except csv.Error:
            # Fall back to counting common delimiters
            sample = text[:4096]
            commas = sample.count(",")
            tabs = sample.count("\t")
            semicolons = sample.count(";")
            delimiter = "," if commas >= tabs and commas >= semicolons else (
                "\t" if tabs >= semicolons else ";"
            )

        reader = csv.DictReader(StringIO(text), delimiter=delimiter)
        rows = []
        for row in reader:
            rows.append(dict(row))
            if len(rows) >= 100:
                break

        # Count total rows
        total_rows = sum(1 for _ in csv.reader(StringIO(text), delimiter=delimiter)) - 1  # minus header

        return json.dumps({
            "file": str(p),
            "format": "csv",
            "delimiter": delimiter,
            "columns": list(rows[0].keys()) if rows else [],
            "row_count": total_rows,
            "rows_shown": len(rows),
            "truncated": len(rows) < total_rows,
            "data": rows,
        }, indent=2, ensure_ascii=False, default=str)

    # ── XLSX ──
    elif suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            return json.dumps({
                "error": "openpyxl is not installed",
                "hint": "Install with: pip install openpyxl",
                "file": str(p),
            }, indent=2)

        wb = openpyxl.load_workbook(p, data_only=True)
        sheets_info = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_list = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 101), values_only=True):
                rows_list.append(list(row))

            sheets_info.append({
                "name": sheet_name,
                "total_rows": ws.max_row,
                "rows_shown": min(ws.max_row, 101),
                "truncated": ws.max_row > 101,
                "headers": rows_list[0] if rows_list else [],
                "data": rows_list[1:101] if len(rows_list) > 1 else [],
            })

        wb.close()
        return json.dumps({
            "file": str(p),
            "format": "xlsx",
            "sheets": sheets_info,
            "sheet_count": len(sheets_info),
        }, indent=2, ensure_ascii=False, default=str)

    # ── JSON ──
    elif suffix == ".json":
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            return json.dumps({
                "error": f"Invalid JSON: {e}",
                "file": str(p),
            }, indent=2)

        if isinstance(data, list):
            return json.dumps({
                "file": str(p),
                "format": "json",
                "type": "array",
                "item_count": len(data),
                "keys_found": list(data[0].keys()) if data and isinstance(data[0], dict) else None,
                "sample": data[:10],
            }, indent=2, ensure_ascii=False, default=str)
        elif isinstance(data, dict):
            return json.dumps({
                "file": str(p),
                "format": "json",
                "type": "object",
                "top_level_keys": list(data.keys()),
                "size_kb": round(len(json.dumps(data)) / 1024, 1),
                "data": data,
            }, indent=2, ensure_ascii=False, default=str)
        else:
            return json.dumps({
                "file": str(p),
                "format": "json",
                "type": type(data).__name__,
                "value": str(data)[:500],
            }, indent=2)

    else:
        return json.dumps({
            "error": f"Unsupported file format: {suffix}",
            "supported_formats": [".csv", ".xlsx", ".xlsm", ".json"],
            "file": str(p),
        }, indent=2)


async def parse_financial_statement(file_path: str) -> str:
    """Parse a financial statement (CSV/XLSX) and extract line items.

    Detects likely financial columns: dates, amounts, categories, descriptions.
    Returns structured financial data suitable for analysis.
    """
    import re

    p = _resolve_path(file_path)
    if not p.exists():
        return json.dumps({
            "error": f"File not found: {file_path}",
            "resolved_path": str(p),
        }, indent=2)

    suffix = p.suffix.lower()

    if suffix == ".csv":
        try:
            text = p.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = p.read_text(encoding="latin-1")

        reader = csv.DictReader(StringIO(text))
        rows = list(reader)
        columns = list(rows[0].keys()) if rows else []

        # Detect financial columns by name patterns
        date_cols = [c for c in columns if re.search(r'date|period|month|year|quarter', c, re.I)]
        amount_cols = [c for c in columns if re.search(r'amount|value|sum|total|balance|debit|credit|revenue|expense|income|cost|price|fee|tax', c, re.I)]
        category_cols = [c for c in columns if re.search(r'cat|type|class|group|account|department|division', c, re.I)]
        desc_cols = [c for c in columns if re.search(r'desc|note|memo|detail|item|name|title|narrative', c, re.I)]

        return json.dumps({
            "file": str(p),
            "format": "csv",
            "row_count": len(rows),
            "columns": columns,
            "detected_financial_columns": {
                "dates": date_cols,
                "amounts": amount_cols,
                "categories": category_cols,
                "descriptions": desc_cols,
            },
            "sample_rows": rows[:20],
        }, indent=2, ensure_ascii=False, default=str)

    elif suffix in (".xlsx", ".xlsm"):
        # Reuse spreadsheet parser and augment
        result = await parse_spreadsheet(file_path)
        base = json.loads(result)
        base["analysis_type"] = "financial_statement"
        return json.dumps(base, indent=2, ensure_ascii=False)

    return await parse_spreadsheet(file_path)


async def parse_api_contract(file_path: str) -> str:
    """Parse an OpenAPI/Swagger spec or Postman collection.

    Extracts: title, version, endpoints, auth requirements, schemas.
    Supports: .json, .yaml, .yml
    """
    import re

    p = _resolve_path(file_path)
    if not p.exists():
        return json.dumps({
            "error": f"File not found: {file_path}",
            "resolved_path": str(p),
        }, indent=2)

    suffix = p.suffix.lower()

    if suffix in (".yaml", ".yml"):
        try:
            import yaml
            data = yaml.safe_load(p.read_text())
        except ImportError:
            return json.dumps({
                "error": "PyYAML is not installed",
                "hint": "Install with: pip install pyyaml",
                "file": str(p),
            }, indent=2)
    elif suffix == ".json":
        data = json.loads(p.read_text())
    else:
        return json.dumps({
            "error": f"Unsupported format: {suffix} — expected .json, .yaml, or .yml",
        }, indent=2)

    # Detect OpenAPI vs Postman
    if "openapi" in data or "swagger" in data:
        # OpenAPI spec
        endpoints = []
        paths = data.get("paths", {})
        for path, methods in paths.items():
            for method, details in methods.items():
                if method in ("get", "post", "put", "patch", "delete", "options", "head"):
                    endpoints.append({
                        "method": method.upper(),
                        "path": path,
                        "summary": details.get("summary", "") or details.get("operationId", ""),
                        "tags": details.get("tags", []),
                        "parameters": [
                            {"name": p.get("name"), "in": p.get("in"), "required": p.get("required")}
                            for p in details.get("parameters", [])
                        ],
                    })

        security = data.get("security", [])
        components_security = data.get("components", {}).get("securitySchemes", {})
        auth_types = list(components_security.keys())

        schemas = list(data.get("components", {}).get("schemas", {}).keys())

        return json.dumps({
            "file": str(p),
            "spec_type": "openapi",
            "info": data.get("info", {}),
            "endpoint_count": len(endpoints),
            "auth_schemes": auth_types,
            "endpoints": endpoints,
            "schema_names": schemas,
        }, indent=2, ensure_ascii=False)

    elif "item" in data and "info" in data:
        # Postman collection
        items = data.get("item", [])

        def extract_requests(items_list, prefix=""):
            requests = []
            for item in items_list:
                if "request" in item:
                    req = item["request"]
                    requests.append({
                        "name": item.get("name"),
                        "method": req.get("method"),
                        "url": req.get("url", {}).get("raw", "") if isinstance(req.get("url"), dict) else str(req.get("url", "")),
                    })
                if "item" in item:
                    requests.extend(extract_requests(item["item"], prefix + item.get("name", "") + " > "))
            return requests

        all_requests = extract_requests(items)
        return json.dumps({
            "file": str(p),
            "spec_type": "postman",
            "info": data.get("info", {}),
            "request_count": len(all_requests),
            "requests": all_requests,
        }, indent=2, ensure_ascii=False)

    else:
        return json.dumps({
            "error": "File parsed but neither OpenAPI nor Postman format detected",
            "top_level_keys": list(data.keys())[:20],
            "file": str(p),
        }, indent=2)


# ── TOOL EXPORT ──────────────────────────────────────────────────────

DATA_PARSER_TOOLS = [
    parse_spreadsheet,
    parse_financial_statement,
    parse_api_contract,
]

__all__ = [
    "DATA_PARSER_TOOLS",
    "parse_spreadsheet",
    "parse_financial_statement",
    "parse_api_contract",
]
